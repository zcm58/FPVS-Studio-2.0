"""Runtime launcher boundary tests."""

from __future__ import annotations

import csv
import os
from pathlib import Path

import pytest
from openpyxl import load_workbook
from tests.unit.runtime_launcher_helpers import (
    PARTICIPANT_NUMBER,
    StubEngine,
    _read_csv_rows,
)

from fpvs_studio.core.compiler import compile_session_plan
from fpvs_studio.core.execution import ParticipantMetadata, SessionExecutionSummary
from fpvs_studio.core.serialization import read_json_file, write_json_file
from fpvs_studio.engines.registry import register_engine, unregister_engine
from fpvs_studio.runtime.export_modes import EXPORT_MODE_COMPACT
from fpvs_studio.runtime.launcher import (
    LaunchSettings,
    launch_session,
)
from fpvs_studio.runtime.session_export import (
    GROUP_SUMMARY_XLSX_SHEET_NAME,
    SESSION_CONDITION_HISTORY_HEADER,
    refresh_participant_summary_if_stale,
    write_group_summary,
    write_participant_summary,
)
from fpvs_studio.runtime.windows_display import WindowsDisplayMode


@pytest.fixture(autouse=True)
def _stable_windows_display_mode(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(
        "fpvs_studio.runtime.display_refresh.query_primary_windows_display_mode",
        lambda: WindowsDisplayMode(r"\\.\DISPLAY1", 60, 1),
    )


def test_session_export_captures_seed_and_runtime_logs(
    multi_condition_project,
    multi_condition_project_root,
) -> None:
    multi_condition_project.settings.fixation_task.accuracy_task_enabled = True
    captures: dict[str, object] = {}
    register_engine("stub-export", lambda: StubEngine(captures))
    try:
        session_plan = compile_session_plan(
            multi_condition_project,
            refresh_hz=60.0,
            project_root=multi_condition_project_root,
            random_seed=77,
        )

        summary = launch_session(
            multi_condition_project_root,
            session_plan,
            participant_number=PARTICIPANT_NUMBER,
            participant_metadata=ParticipantMetadata(
                age=72,
                sex="Female",
                handedness="Right handed",
                colorblind=True,
            ),
            launch_settings=LaunchSettings(engine_name="stub-export"),
        )
    finally:
        unregister_engine("stub-export")

    assert summary.output_dir is not None
    session_output_dir = multi_condition_project_root / Path(summary.output_dir)
    exported_summary = read_json_file(
        session_output_dir / "session_summary.json",
        SessionExecutionSummary,
    )

    assert exported_summary.random_seed == 77
    assert exported_summary.participant_number == PARTICIPANT_NUMBER
    assert exported_summary.participant_metadata == ParticipantMetadata(
        age=72,
        sex="Female",
        handedness="Right handed",
        colorblind=True,
    )
    assert exported_summary.realized_block_orders == [
        block.condition_order for block in session_plan.blocks
    ]
    assert (session_output_dir / "frame_intervals.csv").is_file()
    assert (session_output_dir / "fixation_events.csv").is_file()
    assert (session_output_dir / "responses.csv").is_file()
    assert (session_output_dir / "trigger_log.csv").is_file()
    assert (multi_condition_project_root / "logs" / "session_condition_history.csv").is_file()
    assert (multi_condition_project_root / "logs" / "participant_summary.csv").is_file()
    assert (multi_condition_project_root / "logs" / "participant_summary.xlsx").is_file()
    assert [run_result.run_id for run_result in exported_summary.run_results] == [
        entry.run_id for entry in session_plan.ordered_entries()
    ]
    assert all(
        run_result.participant_number == PARTICIPANT_NUMBER
        for run_result in exported_summary.run_results
    )

    conditions_rows = _read_csv_rows(session_output_dir / "conditions.csv")
    event_rows = _read_csv_rows(session_output_dir / "events.csv")
    fixation_rows = _read_csv_rows(session_output_dir / "fixation_events.csv")
    response_rows = _read_csv_rows(session_output_dir / "responses.csv")
    trigger_rows = _read_csv_rows(session_output_dir / "trigger_log.csv")
    participant_metadata_rows = _read_csv_rows(session_output_dir / "participant_metadata.csv")
    condition_history_rows = _read_csv_rows(
        multi_condition_project_root / "logs" / "session_condition_history.csv"
    )
    participant_summary_rows = _read_csv_rows(
        multi_condition_project_root / "logs" / "participant_summary.csv"
    )

    assert [row["run_id"] for row in conditions_rows] == [
        entry.run_id for entry in session_plan.ordered_entries()
    ]
    assert [row["run_id"] for row in condition_history_rows] == [
        entry.run_id for entry in session_plan.ordered_entries()
    ]
    assert [row["run_seed"] for row in condition_history_rows] == [
        str(entry.run_spec.random_seed) for entry in session_plan.ordered_entries()
    ]
    assert event_rows
    assert {
        "stimulus_modality",
        "stimulus_id",
        "stimulus_value",
        "image_path",
        "text",
    }.issubset(event_rows[0])
    assert event_rows[0]["stimulus_modality"] == "image"
    assert event_rows[0]["stimulus_value"] == event_rows[0]["image_path"]
    assert event_rows[0]["text"] == ""
    assert all(row["participant_number"] == PARTICIPANT_NUMBER for row in condition_history_rows)
    assert participant_metadata_rows == [
        {
            "participant_number": PARTICIPANT_NUMBER,
            "participant_age": "72",
            "participant_sex": "Female",
            "participant_handedness": "Right handed",
            "participant_colorblind": "Yes",
        }
    ]
    assert all(row["participant_age"] == "72" for row in condition_history_rows)
    assert all(row["participant_sex"] == "Female" for row in condition_history_rows)
    assert all(
        row["participant_handedness"] == "Right handed"
        for row in condition_history_rows
    )
    assert all(row["participant_colorblind"] == "Yes" for row in condition_history_rows)
    assert all(row["session_seed"] == "77" for row in condition_history_rows)
    assert all(row["output_dir"] == summary.output_dir for row in condition_history_rows)
    assert all(row["block_accuracy_percent"] for row in condition_history_rows)
    fixation_summaries = [
        run_result.fixation_task_summary for run_result in exported_summary.run_results
    ]
    assert all(fixation_summary is not None for fixation_summary in fixation_summaries)
    total_targets = sum(
        fixation_summary.total_targets
        for fixation_summary in fixation_summaries
        if fixation_summary is not None
    )
    hit_count = sum(
        fixation_summary.hit_count
        for fixation_summary in fixation_summaries
        if fixation_summary is not None
    )
    false_alarm_count = sum(
        fixation_summary.false_alarm_count
        for fixation_summary in fixation_summaries
        if fixation_summary is not None
    )
    weighted_rt_ms = (
        sum(
            fixation_summary.mean_rt_ms * fixation_summary.hit_count
            for fixation_summary in fixation_summaries
            if fixation_summary is not None and fixation_summary.mean_rt_ms is not None
        )
        / hit_count
    )
    participant_summary_row = participant_summary_rows[0]
    assert participant_summary_rows == [participant_summary_row]
    assert participant_summary_row["PID"] == PARTICIPANT_NUMBER
    assert participant_summary_row["Age"] == "72"
    assert participant_summary_row["Sex"] == "Female"
    assert participant_summary_row["Handedness"] == "Right handed"
    assert participant_summary_row["Colorblind"] == "Yes"
    assert participant_summary_row["Session ID"] == session_plan.session_id
    assert participant_summary_row["Condition Display Order Seed"] == "77"
    assert participant_summary_row["Image Display Order Seeds"] == "; ".join(
        f"{entry.run_id}={entry.run_spec.random_seed}"
        for entry in session_plan.ordered_entries()
    )
    assert participant_summary_row["Total Targets"] == str(total_targets)
    assert participant_summary_row["Hits"] == str(hit_count)
    assert participant_summary_row["False Alarms"] == str(false_alarm_count)
    assert participant_summary_row["Aborted Y/N"] == "N"
    assert participant_summary_row["Include In Analysis"] == "Y"
    assert participant_summary_row["Mean Accuracy Across All Conditions (%)"] == (
        f"{(hit_count / total_targets) * 100.0:.2f}"
    )
    assert participant_summary_row[
        "Mean Reaction Time Across All Conditions (ms)"
    ] == f"{weighted_rt_ms:.2f}"
    assert len(fixation_rows) == sum(
        len(run_result.fixation_responses) for run_result in exported_summary.run_results
    )
    assert len(response_rows) == sum(
        len(run_result.response_log) for run_result in exported_summary.run_results
    )
    assert len(trigger_rows) == sum(
        len(run_result.trigger_log) for run_result in exported_summary.run_results
    )

    workbook = load_workbook(multi_condition_project_root / "logs" / "participant_summary.xlsx")
    worksheet = workbook["Participant Summary"]
    assert worksheet.freeze_panes == "A2"
    assert worksheet.auto_filter.ref == worksheet.dimensions
    workbook_header = [cell.value for cell in worksheet[1]]
    assert workbook_header == list(participant_summary_row)
    workbook_values = [cell.value for cell in worksheet[2]]
    assert workbook_values[0] == PARTICIPANT_NUMBER
    assert workbook_values[8] == total_targets
    assert workbook_values[9] == hit_count
    assert workbook_values[10] == false_alarm_count
    assert workbook_values[11] == "N"
    assert workbook_values[12] == "Y"
    assert workbook_values[13] == float(participant_summary_row[
        "Mean Accuracy Across All Conditions (%)"
    ])
    assert workbook_values[14] == float(participant_summary_row[
        "Mean Reaction Time Across All Conditions (ms)"
    ])
    for row in worksheet.iter_rows():
        for cell in row:
            assert cell.alignment.horizontal == "center"
            assert cell.alignment.vertical == "center"
    for column_index in range(1, worksheet.max_column + 1):
        letter = worksheet.cell(row=1, column=column_index).column_letter
        expected_width = min(
            max(
                len(str(worksheet.cell(row=row_index, column=column_index).value))
                for row_index in range(1, worksheet.max_row + 1)
                if worksheet.cell(row=row_index, column=column_index).value is not None
            )
            + 2,
            255,
        )
        assert worksheet.column_dimensions[letter].width == expected_width


def test_compact_session_export_updates_summary_logs_without_runs_folder(
    multi_condition_project,
    multi_condition_project_root,
) -> None:
    multi_condition_project.settings.fixation_task.accuracy_task_enabled = True
    captures: dict[str, object] = {}
    register_engine("stub-compact-export", lambda: StubEngine(captures))
    try:
        session_plan = compile_session_plan(
            multi_condition_project,
            refresh_hz=60.0,
            project_root=multi_condition_project_root,
            random_seed=78,
        )

        summary = launch_session(
            multi_condition_project_root,
            session_plan,
            participant_number=PARTICIPANT_NUMBER,
            participant_metadata=ParticipantMetadata(
                age=73,
                sex="Male",
                handedness="Ambidextrous",
                colorblind=False,
            ),
            launch_settings=LaunchSettings(
                engine_name="stub-compact-export",
                export_mode=EXPORT_MODE_COMPACT,
            ),
        )
    finally:
        unregister_engine("stub-compact-export")

    assert summary.aborted is False
    assert summary.output_dir is None
    assert all(run_result.output_dir is None for run_result in summary.run_results)
    assert not (multi_condition_project_root / "runs").exists()
    assert (multi_condition_project_root / "logs" / "session_condition_history.csv").is_file()
    assert (multi_condition_project_root / "logs" / "participant_summary.csv").is_file()
    assert (multi_condition_project_root / "logs" / "participant_summary.xlsx").is_file()

    condition_history_rows = _read_csv_rows(
        multi_condition_project_root / "logs" / "session_condition_history.csv"
    )
    participant_summary_rows = _read_csv_rows(
        multi_condition_project_root / "logs" / "participant_summary.csv"
    )

    assert [row["run_id"] for row in condition_history_rows] == [
        entry.run_id for entry in session_plan.ordered_entries()
    ]
    assert all(row["output_dir"] == "" for row in condition_history_rows)
    assert all(row["participant_number"] == PARTICIPANT_NUMBER for row in condition_history_rows)
    assert all(row["participant_age"] == "73" for row in condition_history_rows)
    assert all(row["participant_sex"] == "Male" for row in condition_history_rows)
    assert all(row["participant_handedness"] == "Ambidextrous" for row in condition_history_rows)
    assert all(row["participant_colorblind"] == "No" for row in condition_history_rows)
    assert all(row["session_seed"] == "78" for row in condition_history_rows)

    participant_summary_row = participant_summary_rows[0]
    assert participant_summary_rows == [participant_summary_row]
    assert participant_summary_row["PID"] == PARTICIPANT_NUMBER
    assert participant_summary_row["Age"] == "73"
    assert participant_summary_row["Sex"] == "Male"
    assert participant_summary_row["Handedness"] == "Ambidextrous"
    assert participant_summary_row["Colorblind"] == "No"
    assert participant_summary_row["Session ID"] == session_plan.session_id
    assert participant_summary_row["Condition Display Order Seed"] == "78"
    assert participant_summary_row["Aborted Y/N"] == "N"
    assert participant_summary_row["Include In Analysis"] == "Y"


def test_participant_summary_backfills_run_seeds_from_session_plan_for_legacy_history(
    multi_condition_project,
    multi_condition_project_root,
) -> None:
    session_plan = compile_session_plan(
        multi_condition_project,
        refresh_hz=60.0,
        project_root=multi_condition_project_root,
        random_seed=77,
    )
    output_dir = "runs/0040"
    write_json_file(multi_condition_project_root / output_dir / "session_plan.json", session_plan)
    history_path = multi_condition_project_root / "logs" / "session_condition_history.csv"
    history_path.parent.mkdir(parents=True, exist_ok=True)
    with history_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=SESSION_CONDITION_HISTORY_HEADER)
        writer.writeheader()
        for entry in session_plan.ordered_entries():
            writer.writerow(
                {
                    "participant_number": "0040",
                    "participant_age": "15",
                    "participant_sex": "Female",
                    "participant_handedness": "Right handed",
                    "session_id": session_plan.session_id,
                    "session_seed": "77",
                    "session_aborted": "False",
                    "output_dir": output_dir,
                    "global_order_index": str(entry.global_order_index + 1),
                    "run_id": entry.run_id,
                    "run_seed": "",
                    "total_targets": "10",
                    "hit_count": "9",
                    "false_alarm_count": "1",
                    "mean_rt_ms": "200.00",
                    "run_aborted": "False",
                }
            )

    summary_path = write_participant_summary(multi_condition_project_root)
    participant_summary_rows = _read_csv_rows(summary_path)

    assert participant_summary_rows[0]["PID"] == "0040"
    assert participant_summary_rows[0]["Condition Display Order Seed"] == "77"
    assert participant_summary_rows[0]["Image Display Order Seeds"] == "; ".join(
        f"{entry.run_id}={entry.run_spec.random_seed}"
        for entry in session_plan.ordered_entries()
    )
    assert participant_summary_rows[0]["Total Targets"] == str(session_plan.total_runs * 10)
    assert participant_summary_rows[0]["Hits"] == str(session_plan.total_runs * 9)
    assert participant_summary_rows[0]["False Alarms"] == str(session_plan.total_runs)
    assert participant_summary_rows[0]["Aborted Y/N"] == "N"
    assert participant_summary_rows[0]["Include In Analysis"] == "Y"
    assert participant_summary_rows[0]["Mean Accuracy Across All Conditions (%)"] == "90.00"
    assert participant_summary_rows[0]["Mean Reaction Time Across All Conditions (ms)"] == "200.00"
    workbook = load_workbook(summary_path.with_suffix(".xlsx"))
    worksheet = workbook["Participant Summary"]
    assert worksheet["A2"].value == "0040"
    assert worksheet["L2"].value == "N"
    assert worksheet["M2"].value == "Y"


def test_participant_summary_excludes_admin_test_participant_ids(tmp_path: Path) -> None:
    history_path = tmp_path / "logs" / "session_condition_history.csv"
    history_path.parent.mkdir(parents=True, exist_ok=True)
    with history_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=SESSION_CONDITION_HISTORY_HEADER)
        writer.writeheader()
        for row in (
            {
                "participant_number": "00",
                "session_id": "session-admin-00",
                "session_seed": "11",
                "session_aborted": "False",
                "output_dir": "runs/P00",
                "global_order_index": "1",
                "run_id": "run-admin-00",
                "run_seed": "101",
                "total_targets": "10",
                "hit_count": "9",
                "false_alarm_count": "0",
                "mean_rt_ms": "300.00",
                "run_aborted": "False",
            },
            {
                "participant_number": "0",
                "session_id": "session-admin-0",
                "session_seed": "22",
                "session_aborted": "False",
                "output_dir": "runs/P0",
                "global_order_index": "1",
                "run_id": "run-admin-0",
                "run_seed": "202",
                "total_targets": "10",
                "hit_count": "8",
                "false_alarm_count": "1",
                "mean_rt_ms": "400.00",
                "run_aborted": "False",
            },
            {
                "participant_number": "1",
                "participant_age": "23",
                "participant_sex": "Male",
                "participant_handedness": "Right handed",
                "session_id": "session-participant-1",
                "session_seed": "33",
                "session_aborted": "False",
                "output_dir": "runs/P1",
                "global_order_index": "1",
                "run_id": "run-participant-1",
                "run_seed": "303",
                "total_targets": "10",
                "hit_count": "10",
                "false_alarm_count": "0",
                "mean_rt_ms": "350.00",
                "run_aborted": "False",
            },
        ):
            writer.writerow(row)

    summary_path = write_participant_summary(tmp_path)
    participant_summary_rows = _read_csv_rows(summary_path)

    assert [row["PID"] for row in participant_summary_rows] == ["1"]
    assert participant_summary_rows[0]["Session ID"] == "session-participant-1"
    assert participant_summary_rows[0]["Image Display Order Seeds"] == (
        "run-participant-1=303"
    )

    workbook = load_workbook(summary_path.with_suffix(".xlsx"))
    worksheet = workbook["Participant Summary"]
    assert worksheet.max_row == 2
    assert worksheet.freeze_panes == "A2"
    assert worksheet.auto_filter.ref == "A1:O2"
    assert worksheet["A2"].value == "1"


def test_refresh_participant_summary_if_stale_rebuilds_missing_xlsx(tmp_path: Path) -> None:
    history_path = tmp_path / "logs" / "session_condition_history.csv"
    history_path.parent.mkdir(parents=True, exist_ok=True)
    with history_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=SESSION_CONDITION_HISTORY_HEADER)
        writer.writeheader()
        writer.writerow(
            {
                "participant_number": "1",
                "session_id": "session-1",
                "session_seed": "11",
                "session_aborted": "False",
                "output_dir": "runs/1",
                "global_order_index": "1",
                "run_id": "run-1",
                "run_seed": "101",
                "total_targets": "10",
                "hit_count": "9",
                "false_alarm_count": "1",
                "mean_rt_ms": "300.00",
                "run_aborted": "False",
            }
        )
    summary_path = write_participant_summary(tmp_path)
    xlsx_path = summary_path.with_suffix(".xlsx")
    xlsx_path.unlink()

    refreshed_path = refresh_participant_summary_if_stale(tmp_path)

    assert refreshed_path == summary_path
    assert xlsx_path.is_file()
    workbook = load_workbook(xlsx_path)
    assert workbook["Participant Summary"]["A2"].value == "1"
    assert refresh_participant_summary_if_stale(tmp_path) is None


def test_refresh_participant_summary_if_stale_regenerates_newer_history(
    tmp_path: Path,
) -> None:
    history_path = tmp_path / "logs" / "session_condition_history.csv"
    history_path.parent.mkdir(parents=True, exist_ok=True)
    with history_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=SESSION_CONDITION_HISTORY_HEADER)
        writer.writeheader()
        writer.writerow(
            {
                "participant_number": "1",
                "session_id": "session-1",
                "session_seed": "11",
                "session_aborted": "False",
                "output_dir": "runs/1",
                "global_order_index": "1",
                "run_id": "run-1",
                "run_seed": "101",
                "total_targets": "10",
                "hit_count": "9",
                "false_alarm_count": "1",
                "mean_rt_ms": "300.00",
                "run_aborted": "False",
            }
        )
    summary_path = write_participant_summary(tmp_path)
    xlsx_path = summary_path.with_suffix(".xlsx")
    with history_path.open("a", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=SESSION_CONDITION_HISTORY_HEADER)
        writer.writerow(
            {
                "participant_number": "2",
                "session_id": "session-2",
                "session_seed": "22",
                "session_aborted": "False",
                "output_dir": "runs/2",
                "global_order_index": "1",
                "run_id": "run-2",
                "run_seed": "202",
                "total_targets": "12",
                "hit_count": "12",
                "false_alarm_count": "0",
                "mean_rt_ms": "250.00",
                "run_aborted": "False",
            }
        )
    old_summary_mtime_ns = max(history_path.stat().st_mtime_ns - 1_000_000, 0)
    os.utime(summary_path, ns=(old_summary_mtime_ns, old_summary_mtime_ns))
    os.utime(xlsx_path, ns=(old_summary_mtime_ns, old_summary_mtime_ns))

    refreshed_path = refresh_participant_summary_if_stale(tmp_path)

    assert refreshed_path == summary_path
    participant_summary_rows = _read_csv_rows(summary_path)
    assert [row["PID"] for row in participant_summary_rows] == ["1", "2"]
    workbook = load_workbook(xlsx_path)
    worksheet = workbook["Participant Summary"]
    assert worksheet.max_row == 3
    assert worksheet["A3"].value == "2"
    assert worksheet["N3"].value == 100
    assert worksheet["O3"].value == 250
    assert refresh_participant_summary_if_stale(tmp_path) is None


def test_group_summary_export_uses_included_sessions_and_weighted_metrics(tmp_path) -> None:
    history_path = tmp_path / "logs" / "session_condition_history.csv"
    history_path.parent.mkdir(parents=True, exist_ok=True)
    with history_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=SESSION_CONDITION_HISTORY_HEADER)
        writer.writeheader()
        for row in (
            {
                "participant_number": "0001",
                "participant_age": "30",
                "participant_sex": "Female",
                "participant_handedness": "Right handed",
                "session_id": "session-included",
                "session_seed": "11",
                "session_aborted": "False",
                "output_dir": "runs/0001",
                "global_order_index": "1",
                "run_id": "run-a",
                "run_seed": "101",
                "total_targets": "10",
                "hit_count": "8",
                "false_alarm_count": "1",
                "mean_rt_ms": "300.00",
                "run_aborted": "False",
            },
            {
                "participant_number": "0001",
                "participant_age": "30",
                "participant_sex": "Female",
                "participant_handedness": "Right handed",
                "session_id": "session-included",
                "session_seed": "11",
                "session_aborted": "False",
                "output_dir": "runs/0001",
                "global_order_index": "2",
                "run_id": "run-b",
                "run_seed": "102",
                "total_targets": "20",
                "hit_count": "18",
                "false_alarm_count": "2",
                "mean_rt_ms": "500.00",
                "run_aborted": "False",
            },
            {
                "participant_number": "0002",
                "participant_age": "40",
                "participant_sex": "Male",
                "participant_handedness": "Left handed",
                "session_id": "session-aborted",
                "session_seed": "22",
                "session_aborted": "True",
                "output_dir": "runs/0002",
                "global_order_index": "1",
                "run_id": "run-c",
                "run_seed": "201",
                "total_targets": "10",
                "hit_count": "10",
                "false_alarm_count": "0",
                "mean_rt_ms": "200.00",
                "run_aborted": "True",
            },
        ):
            writer.writerow(row)

    output_path = write_group_summary(tmp_path, tmp_path / "exports" / "final-summary")

    assert output_path == tmp_path / "exports" / "final-summary.xlsx"
    workbook = load_workbook(output_path)
    worksheet = workbook[GROUP_SUMMARY_XLSX_SHEET_NAME]
    header = [cell.value for cell in worksheet[1]]
    column = {name: index + 1 for index, name in enumerate(header)}
    assert worksheet.freeze_panes == "A2"
    assert worksheet.auto_filter.ref == worksheet.dimensions

    group_row = {
        name: worksheet.cell(row=2, column=index).value
        for name, index in column.items()
    }
    assert group_row["Row Type"] == "Group Summary"
    assert group_row["PID"] == "GROUP_INCLUDED"
    assert group_row["Included Sessions"] == 1
    assert group_row["Excluded Sessions"] == 1
    assert group_row["Total Targets"] == 30
    assert group_row["Hits"] == 26
    assert group_row["False Alarms"] == 3
    assert group_row["Aborted Y/N"] == "N"
    assert group_row["Include In Analysis"] == "Y"
    assert group_row["Mean Accuracy Across All Conditions (%)"] == 86.67
    assert group_row["Mean Reaction Time Across All Conditions (ms)"] == 438.46
    assert str(group_row["Generated At UTC"]).endswith("+00:00")

    participant_types = [
        worksheet.cell(row=row_index, column=column["Row Type"]).value
        for row_index in range(3, worksheet.max_row + 1)
    ]
    participant_pids = [
        worksheet.cell(row=row_index, column=column["PID"]).value
        for row_index in range(3, worksheet.max_row + 1)
    ]
    included_flags = [
        worksheet.cell(row=row_index, column=column["Include In Analysis"]).value
        for row_index in range(3, worksheet.max_row + 1)
    ]
    assert participant_types == ["Participant Session", "Participant Session"]
    assert participant_pids == ["0001", "0002"]
    assert included_flags == ["Y", "N"]
    for row in worksheet.iter_rows():
        for cell in row:
            assert cell.alignment.horizontal == "center"
            assert cell.alignment.vertical == "center"


