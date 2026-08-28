from __future__ import annotations

import csv
from dataclasses import FrozenInstanceError
from pathlib import Path

import pytest
from openpyxl import load_workbook  # type: ignore[import-untyped]

from fpvs_studio.runtime.fixation_report import (
    FIXATION_TASK_ACCURACY_FILENAME,
    FIXATION_TASK_ACCURACY_HEADER,
    FIXATION_TASK_ACCURACY_SHEET_NAME,
    FixationConditionSummary,
    FixationCrossDataSummary,
    FixationDataError,
    FixationExportError,
    load_fixation_cross_data,
    write_fixation_task_accuracy_xlsx,
)
from fpvs_studio.runtime.session_export import SESSION_CONDITION_HISTORY_HEADER


def test_write_fixation_task_accuracy_xlsx_preserves_flat_values_and_default_fills(
    tmp_path: Path,
) -> None:
    summary = _fixation_summary()

    output_path = write_fixation_task_accuracy_xlsx(
        summary,
        tmp_path / "exports" / "fixation-results.xlsx",
    )

    assert output_path == tmp_path / "exports" / "fixation-results.xlsx"
    workbook = load_workbook(output_path)
    assert workbook.sheetnames == [FIXATION_TASK_ACCURACY_SHEET_NAME]
    worksheet = workbook[FIXATION_TASK_ACCURACY_SHEET_NAME]
    assert worksheet.max_row == 4
    assert worksheet.max_column == len(FIXATION_TASK_ACCURACY_HEADER)
    assert [cell.value for cell in worksheet[1]] == list(FIXATION_TASK_ACCURACY_HEADER)
    assert list(worksheet.values)[1:] == [
        ("Overall", None, None, 3, 30, 40, 75, 312.5),
        ("Condition", "condition-a", "Faces", 3, 8, 10, 80, 300),
        ("Condition", "condition-b", "Words", 2, 22, 30, 73.33333333333333, None),
    ]
    assert worksheet.auto_filter.ref == "A1:H4"
    assert worksheet.freeze_panes == "A2"

    numeric_cells = ("D2", "E2", "F2", "G2", "H2", "D3", "E3", "F3", "G3", "H3")
    for coordinate in numeric_cells:
        cell = worksheet[coordinate]
        assert isinstance(cell.value, (int, float))
        assert cell.data_type == "n"

    for coordinate in ("G2", "H2", "G3", "H3", "G4"):
        assert worksheet[coordinate].number_format == "0.0"
    for coordinate in ("D2", "E2", "F2", "D3", "E3", "F3"):
        assert worksheet[coordinate].number_format == "General"

    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            assert cell.alignment.horizontal == "center"
            assert cell.alignment.vertical == "center"
            assert cell.alignment.wrap_text is True
            if isinstance(cell.value, str):
                assert cell.data_type == "s"
            assert cell.fill.fill_type is None
            assert cell.fill.fgColor.rgb == "00000000"
            assert cell.fill.bgColor.rgb == "00000000"


def test_write_fixation_task_accuracy_xlsx_adds_suffix_without_changing_logs(
    tmp_path: Path,
) -> None:
    project_root = tmp_path / "project"
    history_path = project_root / "logs" / "session_condition_history.csv"
    history_path.parent.mkdir(parents=True)
    history_path.write_text("history stays untouched\n", encoding="utf-8")
    original_history = history_path.read_bytes()

    output_path = write_fixation_task_accuracy_xlsx(
        _fixation_summary(),
        project_root / "exports" / "accuracy",
    )

    assert output_path == project_root / "exports" / "accuracy.xlsx"
    assert output_path.is_file()
    assert history_path.read_bytes() == original_history
    assert not (project_root / "logs" / "participant_summary.csv").exists()
    assert not (project_root / "logs" / "participant_summary.xlsx").exists()


def test_write_fixation_task_accuracy_xlsx_appends_after_an_alternate_suffix(
    tmp_path: Path,
) -> None:
    selected_path = tmp_path / "report.csv"
    selected_path.write_text("selected file remains", encoding="utf-8")
    replacement_collision = tmp_path / "report.xlsx"
    replacement_collision.write_text("existing workbook remains", encoding="utf-8")

    output_path = write_fixation_task_accuracy_xlsx(_fixation_summary(), selected_path)

    assert output_path == tmp_path / "report.csv.xlsx"
    assert output_path.is_file()
    assert selected_path.read_text(encoding="utf-8") == "selected file remains"
    assert replacement_collision.read_text(encoding="utf-8") == "existing workbook remains"


def test_write_fixation_task_accuracy_xlsx_keeps_formula_like_names_as_wrapped_text(
    tmp_path: Path,
) -> None:
    condition_id = "=SUM(1, 2)"
    condition_name = "=" + ("A long literal condition name for export verification " * 4)
    summary = FixationCrossDataSummary(
        included_session_count=1,
        total_targets=10,
        hit_count=8,
        accuracy_percent=80.0,
        mean_rt_ms=300.0,
        conditions=(
            FixationConditionSummary(
                condition_id=condition_id,
                condition_name=condition_name,
                included_session_count=1,
                total_targets=10,
                hit_count=8,
                accuracy_percent=80.0,
                mean_rt_ms=300.0,
            ),
        ),
    )
    output_path = write_fixation_task_accuracy_xlsx(summary, tmp_path / "literal-text")

    worksheet = load_workbook(output_path)[FIXATION_TASK_ACCURACY_SHEET_NAME]
    assert worksheet["B3"].value == condition_id
    assert worksheet["B3"].data_type == "s"
    assert worksheet["C3"].value == condition_name
    assert worksheet["C3"].data_type == "s"
    assert worksheet["C3"].alignment.wrap_text is True
    assert worksheet.row_dimensions[3].height is not None
    assert worksheet.row_dimensions[3].height > 15.0
    assert worksheet["C3"].fill.fill_type is None

    data_only_worksheet = load_workbook(
        output_path,
        data_only=True,
    )[FIXATION_TASK_ACCURACY_SHEET_NAME]
    assert data_only_worksheet["B3"].value == condition_id
    assert data_only_worksheet["C3"].value == condition_name


def test_write_fixation_task_accuracy_xlsx_replaces_an_existing_workbook(
    tmp_path: Path,
) -> None:
    output_path = tmp_path / FIXATION_TASK_ACCURACY_FILENAME
    write_fixation_task_accuracy_xlsx(_fixation_summary(), output_path)
    replacement = FixationCrossDataSummary(
        included_session_count=1,
        total_targets=2,
        hit_count=1,
        accuracy_percent=50.0,
        mean_rt_ms=450.0,
        conditions=(),
    )

    returned_path = write_fixation_task_accuracy_xlsx(replacement, output_path)

    assert returned_path == output_path
    worksheet = load_workbook(output_path)[FIXATION_TASK_ACCURACY_SHEET_NAME]
    assert worksheet.max_row == 2
    assert [cell.value for cell in worksheet[2]] == [
        "Overall",
        None,
        None,
        1,
        1,
        2,
        50,
        450,
    ]


def test_write_fixation_task_accuracy_xlsx_wraps_filesystem_errors(tmp_path: Path) -> None:
    blocking_parent = tmp_path / "not-a-directory"
    blocking_parent.write_text("occupied", encoding="utf-8")

    with pytest.raises(FixationExportError, match="could not be saved"):
        write_fixation_task_accuracy_xlsx(
            _fixation_summary(),
            blocking_parent / "accuracy",
        )


def test_load_fixation_cross_data_pools_repeated_runs_and_sessions_by_condition(
    tmp_path: Path,
) -> None:
    project_root = tmp_path / "project"
    _write_history(
        project_root,
        [
            _history_row(
                condition_id="condition-a",
                condition_name="Faces old name",
                total_targets="10",
                hit_count="8",
                mean_rt_ms="300",
            ),
            _history_row(
                condition_id="condition-b",
                condition_name="Shared label",
                total_targets="5",
                hit_count="5",
                mean_rt_ms="200",
            ),
            _history_row(
                condition_id="condition-a",
                condition_name="",
                total_targets="20",
                hit_count="10",
                mean_rt_ms="500",
            ),
            _history_row(
                output_dir="runs/P01_run2",
                condition_id="condition-a",
                condition_name="Faces renamed",
                total_targets="2",
                hit_count="2",
                mean_rt_ms="100",
            ),
            _history_row(
                output_dir="runs/P01_run2",
                condition_id="condition-b",
                condition_name="Shared label",
                total_targets="15",
                hit_count="0",
                mean_rt_ms="",
            ),
            _history_row(
                output_dir="runs/P01_run2",
                condition_id="condition-c",
                condition_name="Shared label",
                total_targets="4",
                hit_count="1",
                mean_rt_ms="600",
            ),
            _history_row(
                condition_id="condition-a",
                condition_name="Faces final name",
            ),
        ],
    )

    summary = load_fixation_cross_data(project_root)

    assert summary is not None
    assert isinstance(summary, FixationCrossDataSummary)
    assert summary.included_session_count == 2
    assert summary.total_targets == 56
    assert summary.hit_count == 26
    assert summary.accuracy_percent == pytest.approx((26 / 56) * 100.0)
    assert summary.mean_rt_ms == pytest.approx(9200 / 26)
    assert isinstance(summary.conditions, tuple)

    condition_a, condition_b, condition_c = summary.conditions
    assert condition_a.condition_id == "condition-a"
    assert condition_a.condition_name == "Faces final name"
    assert condition_a.included_session_count == 2
    assert condition_a.total_targets == 32
    assert condition_a.hit_count == 20
    assert condition_a.accuracy_percent == 62.5
    assert condition_a.mean_rt_ms == 380.0

    assert condition_b.condition_id == "condition-b"
    assert condition_b.included_session_count == 2
    assert condition_b.total_targets == 20
    assert condition_b.hit_count == 5
    assert condition_b.accuracy_percent == 25.0
    assert condition_b.mean_rt_ms == 200.0

    assert condition_c.condition_id == "condition-c"
    assert condition_c.condition_name == "Shared label"
    assert condition_c.included_session_count == 1
    assert condition_c.total_targets == 4
    assert condition_c.hit_count == 1
    assert condition_c.accuracy_percent == 25.0
    assert condition_c.mean_rt_ms == 600.0

    with pytest.raises(FrozenInstanceError):
        summary.total_targets = 0  # type: ignore[misc]


def test_load_fixation_cross_data_excludes_admin_and_whole_aborted_sessions(
    tmp_path: Path,
) -> None:
    project_root = tmp_path / "project"
    _write_history(
        project_root,
        [
            _history_row(
                participant_number="001",
                session_id="included",
                output_dir="runs/P001",
                condition_name="Included name",
                total_targets="10",
                hit_count="7",
                mean_rt_ms="250",
            ),
            _history_row(
                participant_number=" 0 ",
                session_id="admin-0",
                output_dir="runs/P0",
                total_targets="100",
                hit_count="100",
                mean_rt_ms="1",
            ),
            _history_row(
                participant_number="00",
                session_id="admin-00",
                output_dir="runs/P00",
                total_targets="100",
                hit_count="100",
                mean_rt_ms="1",
            ),
            _history_row(
                participant_number="002",
                session_id="run-aborted",
                output_dir="runs/P002",
                condition_name="Excluded rename",
                total_targets="20",
                hit_count="20",
                mean_rt_ms="100",
            ),
            _history_row(
                participant_number="002",
                session_id="run-aborted",
                output_dir="runs/P002",
                condition_id="condition-b",
                condition_name="Aborted row",
                run_aborted="True",
                total_targets="10",
                hit_count="0",
            ),
            _history_row(
                participant_number="003",
                session_id="session-aborted",
                output_dir="runs/P003",
                session_aborted="True",
                total_targets="10",
                hit_count="10",
                mean_rt_ms="100",
            ),
        ],
    )

    summary = load_fixation_cross_data(project_root)

    assert summary is not None
    assert summary.included_session_count == 1
    assert summary.total_targets == 10
    assert summary.hit_count == 7
    assert summary.accuracy_percent == 70.0
    assert summary.mean_rt_ms == 250.0
    assert len(summary.conditions) == 1
    assert summary.conditions[0].condition_name == "Included name"


def test_load_fixation_cross_data_returns_summary_with_no_rt_when_there_are_no_hits(
    tmp_path: Path,
) -> None:
    project_root = tmp_path / "project"
    _write_history(
        project_root,
        [
            _history_row(
                total_targets="12",
                hit_count="0",
                mean_rt_ms="",
            )
        ],
    )

    summary = load_fixation_cross_data(project_root)

    assert summary is not None
    assert summary.accuracy_percent == 0.0
    assert summary.mean_rt_ms is None
    assert summary.conditions[0].mean_rt_ms is None


def test_missing_empty_header_only_and_zero_target_history_are_no_data(
    tmp_path: Path,
) -> None:
    missing_project = tmp_path / "missing"
    assert load_fixation_cross_data(missing_project) is None

    empty_project = tmp_path / "empty"
    empty_path = empty_project / "logs" / "session_condition_history.csv"
    empty_path.parent.mkdir(parents=True)
    empty_path.write_text("", encoding="utf-8")
    assert load_fixation_cross_data(empty_project) is None

    header_project = tmp_path / "header-only"
    _write_history(header_project, [])
    assert load_fixation_cross_data(header_project) is None

    zero_target_project = tmp_path / "zero-target"
    _write_history(zero_target_project, [_history_row()])
    assert load_fixation_cross_data(zero_target_project) is None


def test_load_fixation_cross_data_is_repeatable_and_isolated_to_active_project(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    active_project = tmp_path / "active"
    other_project = tmp_path / "other"
    working_directory = tmp_path / "working-directory"
    _write_history(
        active_project,
        [_history_row(total_targets="4", hit_count="3", mean_rt_ms="240")],
    )
    _write_history(
        other_project,
        [_history_row(total_targets="100", hit_count="0", mean_rt_ms="")],
    )
    _write_history(
        working_directory,
        [_history_row(total_targets="50", hit_count="50", mean_rt_ms="1")],
    )
    active_history = active_project / "logs" / "session_condition_history.csv"
    original_bytes = active_history.read_bytes()
    monkeypatch.chdir(working_directory)

    first = load_fixation_cross_data(active_project)
    second = load_fixation_cross_data(active_project)

    assert first == second
    assert first is not None
    assert first.total_targets == 4
    assert first.hit_count == 3
    assert active_history.read_bytes() == original_bytes
    assert not (active_project / "logs" / "participant_summary.csv").exists()
    assert not (active_project / "logs" / "participant_summary.xlsx").exists()


@pytest.mark.parametrize(
    ("column", "value", "message"),
    [
        ("total_targets", "not-a-number", "invalid total_targets"),
        ("hit_count", "11", "more hits than targets"),
        ("session_aborted", "perhaps", "invalid session_aborted"),
        ("mean_rt_ms", "NaN", "invalid mean_rt_ms"),
    ],
)
def test_malformed_fixation_values_raise_recoverable_error(
    tmp_path: Path,
    column: str,
    value: str,
    message: str,
) -> None:
    project_root = tmp_path / column
    row = _history_row(total_targets="10", hit_count="1", mean_rt_ms="200")
    row[column] = value
    _write_history(project_root, [row])

    with pytest.raises(FixationDataError, match=message):
        load_fixation_cross_data(project_root)


def test_missing_required_history_columns_raise_recoverable_error(tmp_path: Path) -> None:
    project_root = tmp_path / "missing-columns"
    history_path = project_root / "logs" / "session_condition_history.csv"
    history_path.parent.mkdir(parents=True)
    history_path.write_text("participant_number,total_targets\n1,10\n", encoding="utf-8")

    with pytest.raises(FixationDataError, match="missing required columns"):
        load_fixation_cross_data(project_root)


def test_unreadable_history_raises_recoverable_error(tmp_path: Path) -> None:
    project_root = tmp_path / "unreadable"
    history_path = project_root / "logs" / "session_condition_history.csv"
    history_path.mkdir(parents=True)

    with pytest.raises(FixationDataError, match="could not be read"):
        load_fixation_cross_data(project_root)


def _write_history(project_root: Path, rows: list[dict[str, str]]) -> Path:
    history_path = project_root / "logs" / "session_condition_history.csv"
    history_path.parent.mkdir(parents=True, exist_ok=True)
    with history_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=SESSION_CONDITION_HISTORY_HEADER)
        writer.writeheader()
        writer.writerows(rows)
    return history_path


def _history_row(**overrides: str) -> dict[str, str]:
    row = {column: "" for column in SESSION_CONDITION_HISTORY_HEADER}
    row.update(
        {
            "participant_number": "01",
            "session_id": "session-shared",
            "output_dir": "runs/P01",
            "session_aborted": "False",
            "run_aborted": "False",
            "condition_id": "condition-a",
            "condition_name": "Faces",
            "total_targets": "0",
            "hit_count": "0",
            "mean_rt_ms": "",
        }
    )
    row.update(overrides)
    return row


def _fixation_summary() -> FixationCrossDataSummary:
    return FixationCrossDataSummary(
        included_session_count=3,
        total_targets=40,
        hit_count=30,
        accuracy_percent=75.0,
        mean_rt_ms=312.5,
        conditions=(
            FixationConditionSummary(
                condition_id="condition-a",
                condition_name="Faces",
                included_session_count=3,
                total_targets=10,
                hit_count=8,
                accuracy_percent=80.0,
                mean_rt_ms=300.0,
            ),
            FixationConditionSummary(
                condition_id="condition-b",
                condition_name="Words",
                included_session_count=2,
                total_targets=30,
                hit_count=22,
                accuracy_percent=73.33333333333333,
                mean_rt_ms=None,
            ),
        ),
    )
