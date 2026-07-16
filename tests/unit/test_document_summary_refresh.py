"""Document-level participant summary refresh behavior."""

from __future__ import annotations

import csv
import os
from pathlib import Path

from openpyxl import load_workbook

from fpvs_studio.core.compiler import compile_session_plan
from fpvs_studio.core.enums import RunMode
from fpvs_studio.core.execution import SessionExecutionSummary
from fpvs_studio.core.project_service import create_project
from fpvs_studio.gui.document import ProjectDocument
from fpvs_studio.runtime.session_export import (
    SESSION_CONDITION_HISTORY_HEADER,
    write_participant_summary,
)


def test_open_existing_project_defers_stale_participant_summary_refresh(
    tmp_path: Path,
) -> None:
    scaffold = create_project(tmp_path, "Open Refresh Project")
    history_path = scaffold.project_root / "logs" / "session_condition_history.csv"
    _write_history_rows(
        history_path,
        [
            _history_row(
                participant_number="1",
                session_id="session-1",
                output_dir="runs/1",
                run_id="run-1",
            )
        ],
    )
    summary_path = write_participant_summary(scaffold.project_root)
    _append_history_rows(
        history_path,
        [
            _history_row(
                participant_number="2",
                session_id="session-2",
                output_dir="runs/2",
                run_id="run-2",
                hit_count="12",
                total_targets="12",
                mean_rt_ms="250.00",
            )
        ],
    )
    _make_summary_outputs_older_than_history(summary_path, history_path)

    document = ProjectDocument.open_existing(scaffold.project_root)

    assert document.project_root == scaffold.project_root
    rows = _read_summary_rows(summary_path)
    assert [row["PID"] for row in rows] == ["1"]

    document.refresh_participant_summary_if_stale()

    rows = _read_summary_rows(summary_path)
    assert [row["PID"] for row in rows] == ["1", "2"]
    workbook = load_workbook(summary_path.with_suffix(".xlsx"))
    assert workbook["Participant Summary"]["A3"].value == "2"


def test_launch_compiled_session_refreshes_stale_participant_summary(
    multi_condition_project,
    multi_condition_project_root: Path,
    monkeypatch,
) -> None:
    document = ProjectDocument(
        project_root=multi_condition_project_root,
        project=multi_condition_project.model_copy(deep=True),
    )
    session_plan = compile_session_plan(
        multi_condition_project,
        refresh_hz=60.0,
        project_root=multi_condition_project_root,
        random_seed=55,
    )
    history_path = multi_condition_project_root / "logs" / "session_condition_history.csv"
    _write_history_rows(
        history_path,
        [
            _history_row(
                participant_number="1",
                session_id="session-1",
                output_dir="runs/1",
                run_id="run-1",
            )
        ],
    )
    summary_path = write_participant_summary(multi_condition_project_root)

    def _launch_session(project_root_arg, _session_plan, **_kwargs):
        _append_history_rows(
            project_root_arg / "logs" / "session_condition_history.csv",
            [
                _history_row(
                    participant_number="2",
                    session_id="session-2",
                    output_dir="runs/2",
                    run_id="run-2",
                    hit_count="12",
                    total_targets="12",
                    mean_rt_ms="250.00",
                )
            ],
        )
        _make_summary_outputs_older_than_history(summary_path, history_path)
        return SessionExecutionSummary(
            project_id=session_plan.project_id,
            session_id=session_plan.session_id,
            engine_name="stub",
            run_mode=RunMode.SESSION,
            participant_number="2",
            total_condition_count=session_plan.total_runs,
            completed_condition_count=session_plan.total_runs,
        )

    monkeypatch.setattr("fpvs_studio.gui.document.launch_session", _launch_session)

    summary = document.launch_compiled_session(
        session_plan,
        participant_number="2",
        display_index=None,
    )

    assert summary.participant_number == "2"
    rows = _read_summary_rows(summary_path)
    assert [row["PID"] for row in rows] == ["1", "2"]
    workbook = load_workbook(summary_path.with_suffix(".xlsx"))
    assert workbook["Participant Summary"]["A3"].value == "2"


def _write_history_rows(path: Path, rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=SESSION_CONDITION_HISTORY_HEADER)
        writer.writeheader()
        writer.writerows(rows)


def _append_history_rows(path: Path, rows: list[dict[str, str]]) -> None:
    with path.open("a", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=SESSION_CONDITION_HISTORY_HEADER)
        writer.writerows(rows)


def _history_row(
    *,
    participant_number: str,
    session_id: str,
    output_dir: str,
    run_id: str,
    hit_count: str = "9",
    total_targets: str = "10",
    mean_rt_ms: str = "300.00",
) -> dict[str, str]:
    return {
        "participant_number": participant_number,
        "session_id": session_id,
        "session_seed": "11",
        "session_aborted": "False",
        "output_dir": output_dir,
        "global_order_index": "1",
        "run_id": run_id,
        "run_seed": "101",
        "total_targets": total_targets,
        "hit_count": hit_count,
        "false_alarm_count": "1",
        "mean_rt_ms": mean_rt_ms,
        "run_aborted": "False",
    }


def _make_summary_outputs_older_than_history(summary_path: Path, history_path: Path) -> None:
    old_summary_mtime_ns = max(history_path.stat().st_mtime_ns - 1_000_000, 0)
    os.utime(summary_path, ns=(old_summary_mtime_ns, old_summary_mtime_ns))
    os.utime(summary_path.with_suffix(".xlsx"), ns=(old_summary_mtime_ns, old_summary_mtime_ns))


def _read_summary_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle))
