"""GUI-neutral fixation-task reporting and explicit workbook export.

The query resolves the canonical project log path, applies the participant-summary
session inclusion rules, and returns immutable aggregation results. The writer exports
those results only to its caller-selected path. Neither operation changes project logs.
"""

from __future__ import annotations

import csv
import math
from collections.abc import Sequence
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook  # type: ignore[import-untyped]
from openpyxl.styles import Alignment  # type: ignore[import-untyped]
from openpyxl.utils import get_column_letter  # type: ignore[import-untyped]

from fpvs_studio.core.paths import logs_dir
from fpvs_studio.runtime.session_export import (
    SESSION_CONDITION_HISTORY_FILENAME,
    _participant_session_history_groups,
    _participant_session_history_identity,
    _session_history_rows_aborted,
    _weighted_mean_rt_ms,
)

__all__ = [
    "FIXATION_TASK_ACCURACY_FILENAME",
    "FIXATION_TASK_ACCURACY_HEADER",
    "FIXATION_TASK_ACCURACY_SHEET_NAME",
    "FixationConditionSummary",
    "FixationCrossDataSummary",
    "FixationDataError",
    "FixationExportError",
    "load_fixation_cross_data",
    "write_fixation_task_accuracy_xlsx",
]

FIXATION_TASK_ACCURACY_FILENAME = "fixation_task_accuracy.xlsx"
FIXATION_TASK_ACCURACY_SHEET_NAME = "Fixation Task Accuracy"
FIXATION_TASK_ACCURACY_HEADER = (
    "Row Type",
    "Condition ID",
    "Condition",
    "Included Sessions",
    "Hits",
    "Targets",
    "Accuracy (%)",
    "Mean Reaction Time (ms)",
)

_REQUIRED_HISTORY_COLUMNS = frozenset(
    {
        "participant_number",
        "session_id",
        "output_dir",
        "session_aborted",
        "run_aborted",
        "condition_id",
        "condition_name",
        "total_targets",
        "hit_count",
        "mean_rt_ms",
    }
)
_TRUE_VALUES = frozenset({"1", "true", "yes", "y"})
_FALSE_VALUES = frozenset({"", "0", "false", "no", "n"})
_CONDITION_NAME_COLUMN_INDEX = 3
_CONDITION_NAME_WRAP_WIDTH = 60
_BASE_EXCEL_ROW_HEIGHT = 15.0
_MAX_WRAPPED_ROW_LINES = 6


class FixationDataError(RuntimeError):
    """Raised when the active project's fixation history cannot be safely reported."""


class FixationExportError(RuntimeError):
    """Raised when a fixation-task accuracy workbook cannot be written."""


@dataclass(frozen=True, slots=True)
class FixationConditionSummary:
    """Pooled fixation metrics for one stable condition id."""

    condition_id: str
    condition_name: str
    included_session_count: int
    total_targets: int
    hit_count: int
    accuracy_percent: float
    mean_rt_ms: float | None


@dataclass(frozen=True, slots=True)
class FixationCrossDataSummary:
    """Pooled fixation metrics across all included participant sessions."""

    included_session_count: int
    total_targets: int
    hit_count: int
    accuracy_percent: float
    mean_rt_ms: float | None
    conditions: tuple[FixationConditionSummary, ...]


def write_fixation_task_accuracy_xlsx(
    summary: FixationCrossDataSummary,
    output_path: Path,
) -> Path:
    """Write ``summary`` as a flat, machine-readable Excel table.

    The caller owns output-path selection. This function writes only the selected
    workbook and never reads or updates project logs.
    """

    path = _fixation_accuracy_xlsx_path(output_path)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = FIXATION_TASK_ACCURACY_SHEET_NAME
    worksheet.append(FIXATION_TASK_ACCURACY_HEADER)
    worksheet.append(_fixation_accuracy_overall_row(summary))
    for condition in summary.conditions:
        worksheet.append(_fixation_accuracy_condition_row(condition))

    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.freeze_panes = "A2"
    _format_fixation_accuracy_worksheet(worksheet)

    try:
        path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(path)
    except OSError as exc:
        raise FixationExportError(
            "The fixation task accuracy workbook could not be saved."
        ) from exc
    return path


def _fixation_accuracy_overall_row(
    summary: FixationCrossDataSummary,
) -> tuple[object, ...]:
    return (
        "Overall",
        None,
        None,
        summary.included_session_count,
        summary.hit_count,
        summary.total_targets,
        summary.accuracy_percent,
        summary.mean_rt_ms,
    )


def _fixation_accuracy_condition_row(
    summary: FixationConditionSummary,
) -> tuple[object, ...]:
    return (
        "Condition",
        summary.condition_id,
        summary.condition_name,
        summary.included_session_count,
        summary.hit_count,
        summary.total_targets,
        summary.accuracy_percent,
        summary.mean_rt_ms,
    )


def _format_fixation_accuracy_worksheet(worksheet: Any) -> None:
    centered = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value is not None:
                cell.alignment = centered
            if isinstance(cell.value, str):
                cell.data_type = "s"
            if cell.row > 1 and cell.column in {7, 8} and isinstance(
                cell.value, (int, float)
            ):
                cell.number_format = "0.0"

    for column_index, cells in enumerate(worksheet.iter_cols(), start=1):
        max_text_length = max(
            (len(str(cell.value)) for cell in cells if cell.value is not None),
            default=0,
        )
        worksheet.column_dimensions[get_column_letter(column_index)].width = min(
            max(max_text_length + 2, 12),
            60,
        )

    for row_index in range(3, worksheet.max_row + 1):
        condition_name = worksheet.cell(
            row=row_index,
            column=_CONDITION_NAME_COLUMN_INDEX,
        ).value
        if not isinstance(condition_name, str):
            continue
        wrapped_line_count = sum(
            max(1, math.ceil(len(line) / _CONDITION_NAME_WRAP_WIDTH))
            for line in (condition_name.splitlines() or [""])
        )
        if wrapped_line_count > 1:
            worksheet.row_dimensions[row_index].height = _BASE_EXCEL_ROW_HEIGHT * min(
                wrapped_line_count,
                _MAX_WRAPPED_ROW_LINES,
            )


def _fixation_accuracy_xlsx_path(output_path: Path) -> Path:
    if output_path.suffix.lower() == ".xlsx":
        return output_path
    return output_path.parent / f"{output_path.name}.xlsx"


def load_fixation_cross_data(project_root: Path) -> FixationCrossDataSummary | None:
    """Load pooled fixation data for ``project_root`` without changing project files.

    Missing, empty, or zero-target history is a normal no-data result. An existing
    history that cannot be read or whose reporting fields are malformed raises
    :class:`FixationDataError` so callers can present a recoverable error state.
    """

    history_path = logs_dir(project_root) / SESSION_CONDITION_HISTORY_FILENAME
    history_rows = _read_history_rows(history_path)
    if not history_rows:
        return None

    _validate_history_rows(history_rows)
    session_groups = _participant_session_history_groups(history_rows)

    included_rows: list[dict[str, str]] = []
    contributing_sessions: set[tuple[str, str, str]] = set()
    condition_rows: dict[str, list[dict[str, str]]] = {}
    condition_sessions: dict[str, set[tuple[str, str, str]]] = {}
    latest_condition_names: dict[str, str] = {}

    included_session_identities = {
        session_identity
        for session_identity, session_rows in session_groups.items()
        if not _session_history_rows_aborted(session_rows)
    }
    for row in history_rows:
        session_identity = _participant_session_history_identity(row)
        if session_identity not in included_session_identities:
            continue
        condition_id = row["condition_id"].strip()
        total_targets = _history_count(row, "total_targets")
        condition_name = row["condition_name"].strip()
        if condition_id and condition_name:
            latest_condition_names[condition_id] = condition_name
        if total_targets <= 0:
            continue

        contributing_sessions.add(session_identity)
        included_rows.append(row)
        condition_rows.setdefault(condition_id, []).append(row)
        condition_sessions.setdefault(condition_id, set()).add(session_identity)

    total_targets = sum(_history_count(row, "total_targets") for row in included_rows)
    if total_targets <= 0:
        return None

    hit_count = sum(_history_count(row, "hit_count") for row in included_rows)
    conditions = tuple(
        _condition_summary(
            condition_id,
            rows,
            session_count=len(condition_sessions[condition_id]),
            condition_name=latest_condition_names.get(condition_id, condition_id),
        )
        for condition_id, rows in condition_rows.items()
    )
    return FixationCrossDataSummary(
        included_session_count=len(contributing_sessions),
        total_targets=total_targets,
        hit_count=hit_count,
        accuracy_percent=(hit_count / total_targets) * 100.0,
        mean_rt_ms=_weighted_mean_rt_ms(included_rows),
        conditions=conditions,
    )


def _condition_summary(
    condition_id: str,
    rows: list[dict[str, str]],
    *,
    session_count: int,
    condition_name: str,
) -> FixationConditionSummary:
    total_targets = sum(_history_count(row, "total_targets") for row in rows)
    hit_count = sum(_history_count(row, "hit_count") for row in rows)
    return FixationConditionSummary(
        condition_id=condition_id,
        condition_name=condition_name,
        included_session_count=session_count,
        total_targets=total_targets,
        hit_count=hit_count,
        accuracy_percent=(hit_count / total_targets) * 100.0,
        mean_rt_ms=_weighted_mean_rt_ms(rows),
    )


def _read_history_rows(path: Path) -> list[dict[str, str]]:
    try:
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle, strict=True)
            fieldnames = reader.fieldnames
            if fieldnames is None:
                return []
            _validate_history_header(fieldnames)

            rows: list[dict[str, str]] = []
            for row_number, row in enumerate(reader, start=2):
                if None in row:
                    raise FixationDataError(
                        f"Fixation history row {row_number} has unexpected extra fields."
                    )
                missing_values = sorted(
                    column for column in _REQUIRED_HISTORY_COLUMNS if row.get(column) is None
                )
                if missing_values:
                    columns = ", ".join(missing_values)
                    raise FixationDataError(
                        f"Fixation history row {row_number} is missing values for: {columns}."
                    )
                rows.append(
                    {
                        str(key): "" if value is None else value
                        for key, value in row.items()
                        if key is not None
                    }
                )
            return rows
    except FileNotFoundError:
        return []
    except FixationDataError:
        raise
    except (OSError, UnicodeError, csv.Error) as exc:
        raise FixationDataError(
            "The active project's fixation history could not be read."
        ) from exc


def _validate_history_header(fieldnames: Sequence[str]) -> None:
    if len(fieldnames) != len(set(fieldnames)):
        raise FixationDataError("Fixation history has duplicate column names.")
    missing_columns = sorted(_REQUIRED_HISTORY_COLUMNS.difference(fieldnames))
    if missing_columns:
        columns = ", ".join(missing_columns)
        raise FixationDataError(
            f"Fixation history is missing required columns: {columns}."
        )


def _validate_history_rows(rows: list[dict[str, str]]) -> None:
    for row_number, row in enumerate(rows, start=2):
        _history_bool(row, "session_aborted", row_number=row_number)
        _history_bool(row, "run_aborted", row_number=row_number)
        total_targets = _history_count(row, "total_targets", row_number=row_number)
        hit_count = _history_count(row, "hit_count", row_number=row_number)
        if hit_count > total_targets:
            raise FixationDataError(
                f"Fixation history row {row_number} has more hits than targets."
            )
        if total_targets > 0 and not row["condition_id"].strip():
            raise FixationDataError(
                f"Fixation history row {row_number} has targets but no condition id."
            )
        mean_rt_ms = _history_float(row, "mean_rt_ms", row_number=row_number)
        if hit_count > 0 and mean_rt_ms is None:
            raise FixationDataError(
                f"Fixation history row {row_number} has hits but no mean reaction time."
            )


def _history_bool(
    row: dict[str, str],
    column: str,
    *,
    row_number: int,
) -> bool:
    value = row[column].strip().lower()
    if value in _TRUE_VALUES:
        return True
    if value in _FALSE_VALUES:
        return False
    raise FixationDataError(
        f"Fixation history row {row_number} has an invalid {column} value."
    )


def _history_count(
    row: dict[str, str],
    column: str,
    *,
    row_number: int | None = None,
) -> int:
    value = row[column].strip()
    if not value:
        return 0
    try:
        count = int(value)
    except ValueError as exc:
        location = f" row {row_number}" if row_number is not None else ""
        raise FixationDataError(
            f"Fixation history{location} has an invalid {column} value."
        ) from exc
    if count < 0:
        location = f" row {row_number}" if row_number is not None else ""
        raise FixationDataError(
            f"Fixation history{location} has a negative {column} value."
        )
    return count


def _history_float(
    row: dict[str, str],
    column: str,
    *,
    row_number: int,
) -> float | None:
    value = row[column].strip()
    if not value:
        return None
    try:
        parsed = float(value)
    except ValueError as exc:
        raise FixationDataError(
            f"Fixation history row {row_number} has an invalid {column} value."
        ) from exc
    if not math.isfinite(parsed) or parsed < 0:
        raise FixationDataError(
            f"Fixation history row {row_number} has an invalid {column} value."
        )
    return parsed
