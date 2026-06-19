"""Writers for runtime-owned neutral session artifacts. It serializes RunSpec, SessionPlan
context, validation reports, and execution summaries into stable JSON, CSV, and XLSX
outputs after playback completes. This module owns export file emission only; scoring,
session flow, and engine behavior are provided by other runtime components."""

from __future__ import annotations

import csv
from collections.abc import Iterable
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook  # type: ignore[import-untyped]
from openpyxl.styles import Alignment, Font, PatternFill  # type: ignore[import-untyped]
from openpyxl.utils import get_column_letter  # type: ignore[import-untyped]

from fpvs_studio.core.enums import DutyCycleMode
from fpvs_studio.core.execution import RunExecutionSummary, SessionExecutionSummary
from fpvs_studio.core.models import DisplayValidationReport, validate_project_relative_path
from fpvs_studio.core.paths import from_project_relative_posix, logs_dir
from fpvs_studio.core.run_spec import RunSpec
from fpvs_studio.core.serialization import read_json_file, write_json_file
from fpvs_studio.core.session_plan import SessionEntry, SessionPlan
from fpvs_studio.core.validation import validate_display_refresh

SESSION_CONDITION_HISTORY_FILENAME = "session_condition_history.csv"
PARTICIPANT_SUMMARY_FILENAME = "participant_summary.csv"
PARTICIPANT_SUMMARY_XLSX_FILENAME = "participant_summary.xlsx"
GROUP_SUMMARY_XLSX_FILENAME = "group_summary.xlsx"
ADMIN_TEST_PARTICIPANT_IDS = frozenset({"0", "00"})
SESSION_CONDITION_HISTORY_HEADER = [
    "logged_at_utc",
    "project_id",
    "project_name",
    "participant_number",
    "participant_age",
    "participant_sex",
    "participant_handedness",
    "session_id",
    "session_seed",
    "session_started_at",
    "session_finished_at",
    "session_aborted",
    "session_abort_reason",
    "output_dir",
    "block_index",
    "index_within_block",
    "global_order_index",
    "condition_id",
    "condition_name",
    "run_id",
    "run_seed",
    "run_started_at",
    "run_finished_at",
    "completed_frames",
    "run_aborted",
    "abort_reason",
    "timing_qc_strict_violation",
    "timing_qc_first_bad_phase",
    "timing_qc_first_bad_frame_index",
    "timing_qc_max_interval_s",
    "timing_qc_strict_violation_reason",
    "total_targets",
    "hit_count",
    "miss_count",
    "false_alarm_count",
    "accuracy_percent",
    "mean_rt_ms",
    "block_accuracy_percent",
]
PARTICIPANT_SUMMARY_HEADER = [
    "PID",
    "Age",
    "Sex",
    "Handedness",
    "Session ID",
    "Condition Display Order Seed",
    "Image Display Order Seeds",
    "Total Targets",
    "Hits",
    "False Alarms",
    "Aborted Y/N",
    "Include In Analysis",
    "Mean Accuracy Across All Conditions (%)",
    "Mean Reaction Time Across All Conditions (ms)",
]
PARTICIPANT_SUMMARY_XLSX_SHEET_NAME = "Participant Summary"
GROUP_SUMMARY_XLSX_SHEET_NAME = "Group Summary"
GROUP_SUMMARY_HEADER = [
    "Row Type",
    "PID",
    "Age",
    "Sex",
    "Handedness",
    "Session ID",
    "Condition Display Order Seed",
    "Image Display Order Seeds",
    "Included Sessions",
    "Excluded Sessions",
    "Total Targets",
    "Hits",
    "False Alarms",
    "Aborted Y/N",
    "Include In Analysis",
    "Mean Accuracy Across All Conditions (%)",
    "Mean Reaction Time Across All Conditions (ms)",
    "Generated At UTC",
]
_PARTICIPANT_SUMMARY_INTEGER_COLUMNS = frozenset(
    {
        "Total Targets",
        "Hits",
        "False Alarms",
    }
)
_PARTICIPANT_SUMMARY_FLOAT_COLUMNS = frozenset(
    {
        "Mean Accuracy Across All Conditions (%)",
        "Mean Reaction Time Across All Conditions (ms)",
    }
)
_GROUP_SUMMARY_INTEGER_COLUMNS = frozenset(
    {
        "Included Sessions",
        "Excluded Sessions",
        *_PARTICIPANT_SUMMARY_INTEGER_COLUMNS,
    }
)
_GROUP_SUMMARY_FLOAT_COLUMNS = _PARTICIPANT_SUMMARY_FLOAT_COLUMNS


def _write_csv(path: Path, header: Iterable[str], rows: Iterable[Iterable[object]] = ()) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(list(header))
        for row in rows:
            writer.writerow(list(row))


def _write_warnings(path: Path, warnings: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(warnings) + ("\n" if warnings else ""), encoding="utf-8")


def _display_report_for_run(run_spec: RunSpec) -> DisplayValidationReport:
    duty_cycle_mode = (
        DutyCycleMode.BLANK_50 if run_spec.display.off_frames > 0 else DutyCycleMode.CONTINUOUS
    )
    return validate_display_refresh(
        run_spec.display.refresh_hz,
        duty_cycle_mode=duty_cycle_mode,
        base_hz=run_spec.condition.base_hz,
    )


def _stimulus_value(event: object) -> str:
    image_path = getattr(event, "image_path", None)
    text = getattr(event, "text", None)
    if image_path is not None:
        return str(image_path)
    if text is not None:
        return str(text)
    return ""


def write_run_artifacts(output_dir: Path, run_spec: RunSpec, summary: RunExecutionSummary) -> None:
    """Write the per-run artifact set for one executed `RunSpec`."""

    output_dir.mkdir(parents=True, exist_ok=True)
    write_json_file(output_dir / "runspec.json", run_spec)
    write_json_file(output_dir / "run_summary.json", summary)
    if summary.runtime_metadata is not None:
        write_json_file(output_dir / "runtime_metadata.json", summary.runtime_metadata)
    write_json_file(output_dir / "display_report.json", _display_report_for_run(run_spec))

    _write_csv(
        output_dir / "events.csv",
        [
            "sequence_index",
            "role",
            "stimulus_modality",
            "stimulus_id",
            "stimulus_value",
            "image_path",
            "text",
            "on_start_frame",
            "on_frames",
            "off_frames",
        ],
        [
            (
                event.sequence_index,
                event.role,
                event.stimulus_modality.value,
                event.stimulus_id,
                _stimulus_value(event),
                event.image_path,
                event.text,
                event.on_start_frame,
                event.on_frames,
                event.off_frames,
            )
            for event in run_spec.stimulus_sequence
        ],
    )
    _write_csv(
        output_dir / "fixation_events.csv",
        [
            "event_index",
            "start_frame",
            "duration_frames",
            "responded",
            "first_response_key",
            "response_frame",
            "response_time_s",
            "rt_frames",
            "outcome",
        ],
        [
            (
                event.event_index,
                event.start_frame,
                event.duration_frames,
                event.responded,
                event.first_response_key,
                event.response_frame,
                event.response_time_s,
                event.rt_frames,
                event.outcome,
            )
            for event in summary.fixation_responses
        ],
    )
    _write_csv(
        output_dir / "responses.csv",
        [
            "response_index",
            "key",
            "frame_index",
            "time_s",
            "matched_event_index",
            "rt_frames",
            "correct",
            "outcome",
        ],
        [
            (
                response.response_index,
                response.key,
                response.frame_index,
                response.time_s,
                response.matched_event_index,
                response.rt_frames,
                response.correct,
                response.outcome,
            )
            for response in summary.response_log
        ],
    )
    _write_csv(
        output_dir / "frame_intervals.csv",
        ["frame_index", "interval_s"],
        [(interval.frame_index, interval.interval_s) for interval in summary.frame_intervals],
    )
    _write_csv(
        output_dir / "trigger_log.csv",
        [
            "trigger_index",
            "frame_index",
            "time_s",
            "code",
            "label",
            "backend_name",
            "status",
            "message",
        ],
        [
            (
                trigger.trigger_index,
                trigger.frame_index,
                trigger.time_s,
                trigger.code,
                trigger.label,
                trigger.backend_name,
                trigger.status,
                trigger.message,
            )
            for trigger in summary.trigger_log
        ],
    )
    _write_warnings(output_dir / "warnings.log", summary.warnings)


def write_session_artifacts(
    output_dir: Path,
    session_plan: SessionPlan,
    summary: SessionExecutionSummary,
    *,
    project_root: Path | None = None,
) -> None:
    """Write the session-level artifact bundle for a compiled/executed session."""

    output_dir.mkdir(parents=True, exist_ok=True)
    write_json_file(output_dir / "session_plan.json", session_plan)
    write_json_file(output_dir / "session_summary.json", summary)
    if summary.runtime_metadata is not None:
        write_json_file(output_dir / "runtime_metadata.json", summary.runtime_metadata)

    _write_csv(
        output_dir / "participant_metadata.csv",
        [
            "participant_number",
            "participant_age",
            "participant_sex",
            "participant_handedness",
        ],
        [
            (
                summary.participant_number or "",
                summary.participant_metadata.age
                if summary.participant_metadata.age is not None
                else "",
                summary.participant_metadata.sex or "",
                summary.participant_metadata.handedness or "",
            )
        ],
    )

    _write_csv(
        output_dir / "conditions.csv",
        [
            "global_order_index",
            "block_index",
            "index_within_block",
            "condition_id",
            "condition_name",
            "run_id",
        ],
        [
            (
                entry.global_order_index,
                entry.block_index,
                entry.index_within_block,
                entry.condition_id,
                entry.condition_name,
                entry.run_id,
            )
            for entry in session_plan.ordered_entries()
        ],
    )
    _write_csv(
        output_dir / "events.csv",
        [
            "run_id",
            "sequence_index",
            "role",
            "stimulus_modality",
            "stimulus_id",
            "stimulus_value",
            "image_path",
            "text",
            "on_start_frame",
            "on_frames",
            "off_frames",
        ],
        [
            (
                entry.run_id,
                event.sequence_index,
                event.role,
                event.stimulus_modality.value,
                event.stimulus_id,
                _stimulus_value(event),
                event.image_path,
                event.text,
                event.on_start_frame,
                event.on_frames,
                event.off_frames,
            )
            for entry in session_plan.ordered_entries()
            for event in entry.run_spec.stimulus_sequence
        ],
    )
    _write_csv(
        output_dir / "fixation_events.csv",
        [
            "run_id",
            "event_index",
            "start_frame",
            "duration_frames",
            "responded",
            "first_response_key",
            "response_frame",
            "response_time_s",
            "rt_frames",
            "outcome",
        ],
        [
            (
                run_result.run_id,
                event.event_index,
                event.start_frame,
                event.duration_frames,
                event.responded,
                event.first_response_key,
                event.response_frame,
                event.response_time_s,
                event.rt_frames,
                event.outcome,
            )
            for run_result in summary.run_results
            for event in run_result.fixation_responses
        ],
    )
    _write_csv(
        output_dir / "responses.csv",
        [
            "run_id",
            "response_index",
            "key",
            "frame_index",
            "time_s",
            "matched_event_index",
            "rt_frames",
            "correct",
            "outcome",
        ],
        [
            (
                run_result.run_id,
                response.response_index,
                response.key,
                response.frame_index,
                response.time_s,
                response.matched_event_index,
                response.rt_frames,
                response.correct,
                response.outcome,
            )
            for run_result in summary.run_results
            for response in run_result.response_log
        ],
    )
    _write_csv(
        output_dir / "frame_intervals.csv",
        ["run_id", "frame_index", "interval_s"],
        [
            (
                run_result.run_id,
                interval.frame_index,
                interval.interval_s,
            )
            for run_result in summary.run_results
            for interval in run_result.frame_intervals
        ],
    )
    _write_csv(
        output_dir / "trigger_log.csv",
        [
            "run_id",
            "trigger_index",
            "frame_index",
            "time_s",
            "code",
            "label",
            "backend_name",
            "status",
            "message",
        ],
        [
            (
                run_result.run_id,
                trigger.trigger_index,
                trigger.frame_index,
                trigger.time_s,
                trigger.code,
                trigger.label,
                trigger.backend_name,
                trigger.status,
                trigger.message,
            )
            for run_result in summary.run_results
            for trigger in run_result.trigger_log
        ],
    )
    _write_warnings(output_dir / "warnings.log", summary.warnings)
    if project_root is not None:
        append_session_condition_history(project_root, session_plan, summary)


def append_session_condition_history(
    project_root: Path,
    session_plan: SessionPlan,
    summary: SessionExecutionSummary,
) -> Path:
    """Append project-level condition-history rows for one launched session."""

    path = logs_dir(project_root) / SESSION_CONDITION_HISTORY_FILENAME
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.is_file() and path.stat().st_size > 0:
        _upgrade_session_condition_history_header(path)
    needs_header = not path.is_file() or path.stat().st_size == 0
    with path.open("a", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        if needs_header:
            writer.writerow(SESSION_CONDITION_HISTORY_HEADER)
        writer.writerows(_session_condition_history_rows(session_plan, summary))
    write_participant_summary(project_root)
    return path


def write_participant_summary(project_root: Path) -> Path:
    """Write compact project-level participant/session summary CSV and XLSX files."""

    path = logs_dir(project_root) / PARTICIPANT_SUMMARY_FILENAME
    xlsx_path = logs_dir(project_root) / PARTICIPANT_SUMMARY_XLSX_FILENAME
    history_path = logs_dir(project_root) / SESSION_CONDITION_HISTORY_FILENAME
    if history_path.is_file() and history_path.stat().st_size > 0:
        _upgrade_session_condition_history_header(history_path)
        history_rows = _read_csv_dict_rows(history_path)
    else:
        history_rows = []
    participant_summary_rows = _participant_summary_rows(project_root, history_rows)
    _write_csv(
        path,
        PARTICIPANT_SUMMARY_HEADER,
        participant_summary_rows,
    )
    _write_participant_summary_xlsx(xlsx_path, participant_summary_rows)
    return path


def refresh_participant_summary_if_stale(project_root: Path) -> Path | None:
    """Regenerate compact participant summaries when history is newer or outputs are missing."""

    log_root = logs_dir(project_root)
    history_path = log_root / SESSION_CONDITION_HISTORY_FILENAME
    if not history_path.is_file() or history_path.stat().st_size == 0:
        return None

    history_mtime_ns = history_path.stat().st_mtime_ns
    summary_paths = (
        log_root / PARTICIPANT_SUMMARY_FILENAME,
        log_root / PARTICIPANT_SUMMARY_XLSX_FILENAME,
    )
    if not any(_summary_output_is_stale(path, history_mtime_ns) for path in summary_paths):
        return None
    return write_participant_summary(project_root)


def _summary_output_is_stale(path: Path, history_mtime_ns: int) -> bool:
    if not path.is_file():
        return True
    return path.stat().st_mtime_ns < history_mtime_ns


def write_group_summary(project_root: Path, output_path: Path) -> Path:
    """Write a manual compact group-level summary workbook for the project."""

    participant_summary_path = write_participant_summary(project_root)
    participant_rows = _read_csv_dict_rows(participant_summary_path)
    if not participant_rows:
        raise ValueError(
            "No participant summary rows are available. Run at least one participant "
            "session before exporting a group summary."
        )

    path = _ensure_xlsx_path(output_path)
    generated_at = datetime.now(timezone.utc).isoformat()
    _write_group_summary_xlsx(path, participant_rows, generated_at=generated_at)
    return path


def _upgrade_session_condition_history_header(path: Path) -> None:
    with path.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle)
        if reader.fieldnames == SESSION_CONDITION_HISTORY_HEADER:
            return
        rows = list(reader)
    _write_csv(
        path,
        SESSION_CONDITION_HISTORY_HEADER,
        ([row.get(column, "") for column in SESSION_CONDITION_HISTORY_HEADER] for row in rows),
    )


def _session_condition_history_rows(
    session_plan: SessionPlan,
    summary: SessionExecutionSummary,
) -> list[list[object]]:
    logged_at = datetime.now(timezone.utc).isoformat()
    run_results_by_id = {run_result.run_id: run_result for run_result in summary.run_results}
    block_accuracy = _block_accuracy_percentages(session_plan, run_results_by_id)
    return [
        _session_condition_history_row(
            entry,
            summary,
            run_results_by_id.get(entry.run_id),
            logged_at=logged_at,
            block_accuracy_percent=block_accuracy.get(entry.block_index),
        )
        for entry in session_plan.ordered_entries()
    ]


def _session_condition_history_row(
    entry: SessionEntry,
    summary: SessionExecutionSummary,
    run_result: RunExecutionSummary | None,
    *,
    logged_at: str,
    block_accuracy_percent: float | None,
) -> list[object]:
    fixation = run_result.fixation_task_summary if run_result is not None else None
    metadata = run_result.runtime_metadata if run_result is not None else None
    return [
        logged_at,
        summary.project_id,
        entry.run_spec.project_name,
        summary.participant_number or "",
        summary.participant_metadata.age
        if summary.participant_metadata.age is not None
        else "",
        summary.participant_metadata.sex or "",
        summary.participant_metadata.handedness or "",
        summary.session_id,
        summary.random_seed if summary.random_seed is not None else "",
        _datetime_value(summary.started_at),
        _datetime_value(summary.finished_at),
        summary.aborted,
        summary.abort_reason or "",
        summary.output_dir or "",
        entry.block_index + 1,
        entry.index_within_block + 1,
        entry.global_order_index + 1,
        entry.condition_id,
        entry.condition_name,
        entry.run_id,
        entry.run_spec.random_seed,
        _datetime_value(run_result.started_at if run_result is not None else None),
        _datetime_value(run_result.finished_at if run_result is not None else None),
        run_result.completed_frames if run_result is not None else "",
        run_result.aborted if run_result is not None else "",
        run_result.abort_reason if run_result is not None and run_result.abort_reason else "",
        metadata.timing_qc_strict_violation if metadata is not None else "",
        metadata.timing_qc_first_bad_phase if metadata is not None else "",
        metadata.timing_qc_first_bad_frame_index
        if metadata is not None and metadata.timing_qc_first_bad_frame_index is not None
        else "",
        _float_value(metadata.timing_qc_max_interval_s if metadata is not None else None),
        metadata.timing_qc_strict_violation_reason
        if metadata is not None and metadata.timing_qc_strict_violation_reason
        else "",
        fixation.total_targets if fixation is not None else "",
        fixation.hit_count if fixation is not None else "",
        fixation.miss_count if fixation is not None else "",
        fixation.false_alarm_count if fixation is not None else "",
        _float_value(fixation.accuracy_percent if fixation is not None else None),
        _float_value(fixation.mean_rt_ms if fixation is not None else None),
        _float_value(block_accuracy_percent),
    ]


def _block_accuracy_percentages(
    session_plan: SessionPlan,
    run_results_by_id: dict[str, RunExecutionSummary],
) -> dict[int, float | None]:
    percentages: dict[int, float | None] = {}
    for block in session_plan.blocks:
        total_targets = 0
        hit_count = 0
        for entry in block.entries:
            run_result = run_results_by_id.get(entry.run_id)
            if run_result is None or run_result.fixation_task_summary is None:
                continue
            fixation = run_result.fixation_task_summary
            total_targets += fixation.total_targets
            hit_count += fixation.hit_count
        percentages[block.block_index] = (
            (hit_count / total_targets) * 100.0 if total_targets > 0 else None
        )
    return percentages


def _read_csv_dict_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle)
        return [
            {str(key): "" if value is None else value for key, value in row.items() if key}
            for row in reader
        ]


def _participant_summary_rows(
    project_root: Path,
    history_rows: list[dict[str, str]],
) -> list[list[object]]:
    grouped_rows: dict[tuple[str, str, str], list[dict[str, str]]] = {}
    for row in history_rows:
        if _is_admin_test_participant_id(row.get("participant_number", "")):
            continue
        key = (
            row.get("participant_number", ""),
            row.get("session_id", ""),
            row.get("output_dir", ""),
        )
        if not any(key):
            continue
        grouped_rows.setdefault(key, []).append(row)
    return [
        _participant_summary_row(project_root, rows)
        for rows in grouped_rows.values()
    ]


def _is_admin_test_participant_id(participant_number: str) -> bool:
    return participant_number.strip() in ADMIN_TEST_PARTICIPANT_IDS


def _participant_summary_row(project_root: Path, rows: list[dict[str, str]]) -> list[object]:
    ordered_rows = sorted(rows, key=_condition_history_order_key)
    total_targets = sum(_csv_int(row.get("total_targets")) or 0 for row in ordered_rows)
    hit_count = sum(_csv_int(row.get("hit_count")) or 0 for row in ordered_rows)
    false_alarm_count = sum(
        _csv_int(row.get("false_alarm_count")) or 0 for row in ordered_rows
    )
    mean_accuracy = (hit_count / total_targets) * 100.0 if total_targets > 0 else None
    mean_rt_ms = _weighted_mean_rt_ms(ordered_rows)
    aborted = any(
        _csv_bool(row.get("session_aborted")) or _csv_bool(row.get("run_aborted"))
        for row in ordered_rows
    )
    return [
        _first_non_blank(ordered_rows, "participant_number"),
        _first_non_blank(ordered_rows, "participant_age"),
        _first_non_blank(ordered_rows, "participant_sex"),
        _first_non_blank(ordered_rows, "participant_handedness"),
        _first_non_blank(ordered_rows, "session_id"),
        _first_non_blank(ordered_rows, "session_seed"),
        _image_display_order_seed_text(project_root, ordered_rows),
        total_targets,
        hit_count,
        false_alarm_count,
        "Y" if aborted else "N",
        "N" if aborted else "Y",
        _float_value(mean_accuracy),
        _float_value(mean_rt_ms),
    ]


def _write_participant_summary_xlsx(path: Path, rows: list[list[object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = PARTICIPANT_SUMMARY_XLSX_SHEET_NAME
    worksheet.append(PARTICIPANT_SUMMARY_HEADER)
    for row in rows:
        worksheet.append(
            [
                _summary_xlsx_value(
                    header,
                    value,
                    integer_columns=_PARTICIPANT_SUMMARY_INTEGER_COLUMNS,
                    float_columns=_PARTICIPANT_SUMMARY_FLOAT_COLUMNS,
                )
                for header, value in zip(PARTICIPANT_SUMMARY_HEADER, row, strict=True)
            ]
        )

    if worksheet.max_column > 0:
        worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.freeze_panes = "A2"
    _format_summary_xlsx(
        worksheet,
        integer_columns=_PARTICIPANT_SUMMARY_INTEGER_COLUMNS,
        float_columns=_PARTICIPANT_SUMMARY_FLOAT_COLUMNS,
    )
    workbook.save(path)


def _write_group_summary_xlsx(
    path: Path,
    participant_rows: list[dict[str, str]],
    *,
    generated_at: str,
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = GROUP_SUMMARY_XLSX_SHEET_NAME
    worksheet.append(GROUP_SUMMARY_HEADER)
    for row in _group_summary_rows(participant_rows, generated_at=generated_at):
        worksheet.append(
            [
                _summary_xlsx_value(
                    header,
                    row.get(header, ""),
                    integer_columns=_GROUP_SUMMARY_INTEGER_COLUMNS,
                    float_columns=_GROUP_SUMMARY_FLOAT_COLUMNS,
                )
                for header in GROUP_SUMMARY_HEADER
            ]
        )

    if worksheet.max_column > 0:
        worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.freeze_panes = "A2"
    _format_summary_xlsx(
        worksheet,
        integer_columns=_GROUP_SUMMARY_INTEGER_COLUMNS,
        float_columns=_GROUP_SUMMARY_FLOAT_COLUMNS,
    )
    workbook.save(path)


def _group_summary_rows(
    participant_rows: list[dict[str, str]],
    *,
    generated_at: str,
) -> list[dict[str, object]]:
    included_rows = [
        row
        for row in participant_rows
        if row.get("Include In Analysis", "").strip().upper() == "Y"
    ]
    total_targets = sum(_csv_int(row.get("Total Targets")) or 0 for row in included_rows)
    hit_count = sum(_csv_int(row.get("Hits")) or 0 for row in included_rows)
    false_alarm_count = sum(_csv_int(row.get("False Alarms")) or 0 for row in included_rows)
    mean_accuracy = (hit_count / total_targets) * 100.0 if total_targets > 0 else None
    mean_rt_ms = _weighted_participant_mean_rt_ms(included_rows)
    included_aborted = any(
        row.get("Aborted Y/N", "").strip().upper() == "Y" for row in included_rows
    )
    group_row: dict[str, object] = {
        "Row Type": "Group Summary",
        "PID": "GROUP_INCLUDED",
        "Included Sessions": len(included_rows),
        "Excluded Sessions": len(participant_rows) - len(included_rows),
        "Total Targets": total_targets,
        "Hits": hit_count,
        "False Alarms": false_alarm_count,
        "Aborted Y/N": "Y" if included_aborted else "N" if included_rows else "",
        "Include In Analysis": "Y" if included_rows else "",
        "Mean Accuracy Across All Conditions (%)": _float_value(mean_accuracy),
        "Mean Reaction Time Across All Conditions (ms)": _float_value(mean_rt_ms),
        "Generated At UTC": generated_at,
    }
    session_rows: list[dict[str, object]] = [
        {
            "Row Type": "Participant Session",
            **{header: row.get(header, "") for header in PARTICIPANT_SUMMARY_HEADER},
            "Generated At UTC": generated_at,
        }
        for row in participant_rows
    ]
    return [group_row, *session_rows]


def _weighted_participant_mean_rt_ms(rows: list[dict[str, str]]) -> float | None:
    weighted_rt_sum = 0.0
    hit_count_for_rt = 0
    for row in rows:
        hit_count = _csv_int(row.get("Hits"))
        mean_rt_ms = _csv_float(row.get("Mean Reaction Time Across All Conditions (ms)"))
        if hit_count is None or hit_count <= 0 or mean_rt_ms is None:
            continue
        weighted_rt_sum += mean_rt_ms * hit_count
        hit_count_for_rt += hit_count
    return weighted_rt_sum / hit_count_for_rt if hit_count_for_rt > 0 else None


def _summary_xlsx_value(
    header: str,
    value: object,
    *,
    integer_columns: frozenset[str],
    float_columns: frozenset[str],
) -> object:
    if header in integer_columns:
        parsed_int = _csv_int(str(value))
        return parsed_int if parsed_int is not None else None
    if header in float_columns:
        parsed_float = _csv_float(str(value))
        return parsed_float if parsed_float is not None else None
    return value


def _format_summary_xlsx(
    worksheet: Any,
    *,
    integer_columns: frozenset[str],
    float_columns: frozenset[str],
) -> None:
    centered = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = centered
            if cell.row == 1:
                cell.font = Font(bold=True)
                cell.fill = header_fill
            if cell.value is None:
                continue
            header = worksheet.cell(row=1, column=cell.column).value
            if header in integer_columns:
                cell.number_format = "0"
            elif header in float_columns:
                cell.number_format = "0.00"

    for column_index, cells in enumerate(worksheet.iter_cols(), start=1):
        max_text_length = max(
            (len(str(cell.value)) for cell in cells if cell.value is not None),
            default=0,
        )
        worksheet.column_dimensions[get_column_letter(column_index)].width = min(
            max_text_length + 2,
            255,
        )


def _ensure_xlsx_path(path: Path) -> Path:
    return path if path.suffix.lower() == ".xlsx" else path.with_suffix(".xlsx")


def _condition_history_order_key(row: dict[str, str]) -> tuple[int, str]:
    return (_csv_int(row.get("global_order_index")) or 0, row.get("run_id", ""))


def _weighted_mean_rt_ms(rows: list[dict[str, str]]) -> float | None:
    weighted_rt_sum = 0.0
    hit_count_for_rt = 0
    for row in rows:
        hit_count = _csv_int(row.get("hit_count"))
        mean_rt_ms = _csv_float(row.get("mean_rt_ms"))
        if hit_count is None or hit_count <= 0 or mean_rt_ms is None:
            continue
        weighted_rt_sum += mean_rt_ms * hit_count
        hit_count_for_rt += hit_count
    return weighted_rt_sum / hit_count_for_rt if hit_count_for_rt > 0 else None


def _image_display_order_seed_text(
    project_root: Path,
    rows: list[dict[str, str]],
) -> str:
    plan_seed_lookup = _session_plan_run_seed_lookup(
        project_root,
        _first_non_blank(rows, "output_dir"),
    )
    parts: list[str] = []
    seen_run_ids: set[str] = set()
    for row in rows:
        run_id = row.get("run_id", "")
        if not run_id or run_id in seen_run_ids:
            continue
        seed = row.get("run_seed", "") or plan_seed_lookup.get(run_id, "")
        if seed:
            parts.append(f"{run_id}={seed}")
        seen_run_ids.add(run_id)
    return "; ".join(parts)


def _session_plan_run_seed_lookup(project_root: Path, output_dir: str) -> dict[str, str]:
    if not output_dir:
        return {}
    try:
        relative_output_dir = validate_project_relative_path(output_dir)
    except ValueError:
        return {}

    session_dir = from_project_relative_posix(project_root, relative_output_dir)
    try:
        session_dir.resolve().relative_to(project_root.resolve())
    except (OSError, ValueError):
        return {}

    session_plan_path = session_dir / "session_plan.json"
    if not session_plan_path.is_file():
        return {}
    try:
        session_plan = read_json_file(session_plan_path, SessionPlan)
    except Exception:
        return {}
    return {
        entry.run_id: str(entry.run_spec.random_seed)
        for entry in session_plan.ordered_entries()
    }


def _first_non_blank(rows: list[dict[str, str]], column: str) -> str:
    for row in rows:
        value = row.get(column, "")
        if value:
            return value
    return ""


def _csv_int(value: str | None) -> int | None:
    if value is None or not value.strip():
        return None
    try:
        return int(value)
    except ValueError:
        return None


def _csv_float(value: str | None) -> float | None:
    if value is None or not value.strip():
        return None
    try:
        return float(value)
    except ValueError:
        return None


def _csv_bool(value: str | None) -> bool:
    return value is not None and value.strip().lower() in {"1", "true", "yes", "y"}


def _datetime_value(value: object | None) -> str:
    if value is None:
        return ""
    return value.isoformat() if hasattr(value, "isoformat") else str(value)


def _float_value(value: float | None) -> str:
    return "" if value is None else f"{value:.2f}"
